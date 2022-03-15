
from tkinter import *
from pathlib import *
import openpyxl as xl
import datetime


root = Tk()
button = Button(root)
e = Entry(root, width=35, borderwidth=5)
e.grid(row=0, column=0, columnspan=3, padx=10, pady=10)


def collect_data(sheet):
    transaction = e.get()
    e.delete(0, END)
    return transaction


def insert_data(sheet):
    row = 5
    counter = 0

    while True:
        counter += 1
        if counter >= row:
            transaction_cell = sheet.cell(counter, 1)
            if transaction_cell.value is None:
                transaction_cell.value = collect_data(sheet)
                break

    # print("Row 5: ", sheet.cell(5, 1).value)
    # print("Row 6: ", sheet.cell(6, 1).value)
    # print("Row 7: ", sheet.cell(7, 1).value)

    # print(tr.value)

    e.delete(0, END)
    wb.save("a.xlsx")


wb = xl.load_workbook("Daily Transactions.xlsx")
sheet = wb['Sheet1']

save_button = Button(root, text='Save', padx=40, pady=20, command=lambda: collect_data(sheet))
save_button.grid(row=1, column=1)

if collect_data(sheet) is True:
    insert_data(sheet)

root.mainloop()
