from tkinter import *
from pathlib import *
import openpyxl as xl
from datetime import *
import time


root = Tk()
button = Button(root)
e = Entry(root, width=35, borderwidth=5)
e.grid(row=0, column=0, columnspan=3, padx=10, pady=10)

wb = xl.load_workbook('Daily Transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a']
#  + " " + str(present_minute) + ":" \+ str(present_second)


def present_date_time():
    present_year = datetime.today().year
    present_month = datetime.today().month
    present_day = datetime.today().day
    present_minute = datetime.today().minute
    present_second = datetime.today().second
    present = str(present_year) + '-' + str(present_month) + "-" + str(present_day)

    return present


def transaction_type_pos(sheet):
    count = 0
    row = 5
    while True:
        count += 1
        if count >= row:
            second_column = sheet.cell(count, 2)
            if second_column.value is None:
                second_column.value = "POS"
                break
    date_time_now = present_date_time()
    filename = date_time_now + '.xlsx'
    wb.save(filename)


def withdrawal(sheet):
    count = 0
    row = 5
    withdraw = e.get()
    while True:
        count += 1
        if count >= row:
            second_column = sheet.cell(count, 4)
            if second_column.value is None:
                second_column.value = withdraw
                break
    date_time_now = present_date_time()
    filename = date_time_now + '.xlsx'
    wb.save(filename)
    e.delete(0, END)


def transaction_type_transfer(sheet):
    count = 0
    row = 5
    while True:
        count += 1
        if count >= row:
            if sheet.cell(count, 2).value is None:
                sheet.cell(count, 2).value = "Transfer"
                break
    date_time_now = present_date_time()
    filename = date_time_now + '.xlsx'
    wb.save(filename)


def input_action(sheet):
    transaction = e.get()
    first_row = 0
    row = 5

    while True:
        first_row += 1
        if first_row >= row:
            if sheet.cell(first_row, 1).value is None:
                sheet.cell(first_row, 1).value = transaction
                each_cell = sheet.cell(first_row, 1).value
                if int(each_cell) < 1600:
                    sheet.cell(first_row, 3).value = 50
                elif int(each_cell) < 5000:
                    sheet.cell(first_row, 3).value = 100
                elif int(each_cell) < 10000:
                    sheet.cell(first_row, 3).value = 200
                elif int(each_cell) < 15000:
                    sheet.cell(first_row, 3).value = 300
                break
        e.delete(0, END)
        date_time_now = present_date_time()
        filename = date_time_now + '.xlsx'
        wb.save(filename)


enter_button = Button(root, text="POS", padx=28, pady=10, command=lambda: transaction_type_pos(sheet))
enter_button.grid(row=1, column=0)

enter_button = Button(root, text="Enter", padx=20, pady=10, command=lambda: input_action(sheet))
enter_button.grid(row=1, column=2)

enter_button = Button(root, text="Transfer", padx=20, pady=10, command=lambda: transaction_type_transfer(sheet))
enter_button.grid(row=2, column=0)

enter_button = Button(root, text="Withdrawal", padx=20, pady=10, command=lambda: withdrawal(sheet))
enter_button.grid(row=2, column=2)

root.mainloop()





#for row in range(5, 8):
#   cell_value = cell_value.value
#    profit_rate = int(Profit(cell_value))
#    print(type(profit_rate))



path = Path("Daily_Transaction")
if path.exists():
    path is TRUE
else:
    print(path.mkdir())

wb.save('work book 86.xslx')








